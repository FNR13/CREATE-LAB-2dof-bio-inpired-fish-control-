# fish_logger/logger.py
import os
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook

class DataLogger:
    """
    A simple logger for recording control loop data and saving to Excel.
    """

    def __init__(self, filename="fish_robot_log.xlsx"):
        self.filename = filename
        self.data = []

    def log(self, t, phi_tail, theta_tail, dynamixel_angle,
            phi_fin, theta_fin):
        """Add one entry to the log."""
        self.data.append({
            "Time (s)": round(t, 4),
            "Tail Phi (°)": round(phi_tail, 3),
            "Tail Theta (°)": round(theta_tail, 3),
            "Tail Real (°)": round(dynamixel_angle, 3),
            "Fin Phi (°)": round(phi_fin, 3),
            "Fin Theta (°)": round(theta_fin, 3),
        })

    def save(self):
        """Save all logged data to an Excel file."""
        if not self.data:
            print("[Logger] No data to save.")
            return

        df = pd.DataFrame(self.data)

        try:
            book = load_workbook(self.filename)
            with pd.ExcelWriter(self.filename, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                start_row = writer.sheets["Sheet1"].max_row
                df.to_excel(writer, index=False, header=False, startrow=start_row)
            print(f"[Logger] Appended {len(df)} rows to {self.filename}")
        except FileNotFoundError:
            df.to_excel(self.filename, index=False)
            print(f"[Logger] Created new log file: {self.filename}")

    def clear(self):
        """Clear the in-memory log buffer."""
        self.data.clear()
        print("[Logger] Log buffer cleared.")

# ---------------------------------------------------------------------

def plot_log(filename="logs/fish_robot_log.xlsx", theta_fin_comparison = False):
    # Load data
    df = pd.read_excel(filename)

    # Adjust Tail Theta by -90°
    df["Fin Theta (°)"] = df["Fin Theta (°)"] - 90

    # Create figure with 3 subplots
    if theta_fin_comparison:
        fig, axes = plt.subplots(3, 1, figsize=(12, 14))
    else:
        fig, axes = plt.subplots(1, 1, figsize=(12, 5))
        axes = [axes]  # Make it iterable

    # ----------------------------------------
    # Subplot 1: Phi comparison (Tail vs Fin)
    # ----------------------------------------
    axes[0].plot(df["Time (s)"], df["Tail Phi (°)"], label="Tail Phi (°)", color="blue")
    axes[0].plot(df["Time (s)"], df["Fin Phi (°)"], label="Fin Phi (°)", color="orange")
    axes[0].set_title("Tail Phi vs Fin Phi over Time")
    axes[0].set_ylabel("Phi (°)")
    axes[0].legend()
    axes[0].grid(True)

    if theta_fin_comparison:
        # ----------------------------------------
        # Subplot 2: Tail Theta vs Phi
        # ----------------------------------------
        axes[1].plot(df["Time (s)"], df["Tail Theta (°)"], label="Tail Theta", color="green")
        axes[1].plot(df["Time (s)"], df["Tail Phi (°)"], label="Tail Phi", color="blue")
        axes[1].set_title("Tail Theta vs Tail Phi over Time")
        axes[1].set_ylabel("Theta (°) / Phi (°)")
        axes[1].legend()
        axes[1].grid(True)

        # ----------------------------------------
        # Subplot 3: Fin Theta vs Phi
        # ----------------------------------------
        axes[2].plot(df["Time (s)"], df["Fin Theta (°)"], label="Fin Theta", color="red")
        axes[2].plot(df["Time (s)"], df["Fin Phi (°)"], label="Fin Phi", color="orange")
        axes[2].set_title("Fin Theta vs Fin Phi over Time")
        axes[2].set_ylabel("Theta (°) / Phi (°)")
        axes[2].legend()
        axes[2].grid(True)

    plt.tight_layout()
    plt.show()


if __name__ == "__main__":
    plot_log("logs/fish_robot_log.xlsx")